import openpyxl
import copy

"""
    openpyxl helps in extracting excel files and fetch data from the sheet
    copy helps in copying the dictionary from the dictionary
"""

#workbook = openpyxl.load_workbook('voting.xlsx')
#sheet = workbook.active
"""In this we have to design and implement several voting rules. Here we have a set of 'n' agents and a set of alternatives 'm'.
Every agent has a preference based on columns in a xlsx file or values in a list, where preference profile is a set of 'n' preference orderings,
one for every agent. Here agents are the row in the xlsx file.
"""



def generatePreferences(values):
    """generatePreferences(values) which inputs a set of numerical values that the agents have for the different alternatives and outputs a preference profile.
    The output (the return) of the generatePreferences function is a dictionary where the keys {1,2,...,n} are the agents 
    and the values are lists that correspond to the preference orderings of those agents. In particular, 
    a value is a list of numbers [x, y , z, ... ] denoting that the agent prefers x to y, y to z and so on. 

    Args:
        values (int): If the valuation of an agent for alternative  is larger than the valuation for alternative , then  should appear before  in the list.
        If the valuation of an agent for alternative  is the same as the valuation for alternative , then  should appear before  in the list if , 
        and otherwise  should appear before  in the list. In other words, in case of ties in the numerical values, 
        alternatives with larger indices are considered to be more preferred by the agent.

    Returns:
        [dictionary]: Holds the dictionary with the sorted and  in the numberical value.
    """
    #values = workbook.active
    #agents and alternatives are colled
    agents=values.max_row
    alternatives=values.max_column
    dict={}
    sortdict={}
    for row in range(1,(agents+1)):
        if row not in dict:
            dict[row]=[]
        for column in range(1,(alternatives+1)):
            dict[row].append (values.cell(row,column).value)
            
    for key,value in dict.items():
        for index,element in enumerate(value):
            sortdict[index + 1]=element
            dict[key]=sorted(sortdict,key=sortdict.get)[::-1]

    return dict
#myprefenence is created to have the value.
#myprefenence=generatePreferences(sheet)
def dictatorship(preferenceProfile, agent):
    """An agent is selected, and the winner is the alternative that this agent ranks first and return of the function should be the winner

    Args:
        preferences (dictionary): The preference holds the dictionary which have the value and keys.
        agent (int): It holds the value and include error handling in case the inputted integer does not correspond to an agent 

    Returns:
        [int]: The final output: the winner is the alternative that this agent ranks first.
    """
      # get the list of highest scorers
    highestScores = []
    preferences=generatePreferences(values)
    maxValue = max(rangedict.values())
    for key, value in rangedict.items():
        if value == maxValue:
            highestScores.append(key)
   
    # find the winner based on the tiebreak
    if tieBreak == 'max':
        return max(highestScores)
    elif tieBreak == 'min':
        return min(highestScores)
    else:
        agent = tieBreak
        temp_dict = {}
        for alternative in highestScores:
            temp_dict[alternative] = preferences[agent].index(alternative)
        # print(preferences[agent])
        # print(temp_dict)
    winner = min(temp_dict, key=temp_dict.get)
    return winner 
    return preferenceProfile[agent][0]
    
def plurality(preferences, tieBreak):
    """ The winner is the alternative that appears the most times in the first position of the agents' preference orderings
    and in case of tie select the single winner.

    Args:
        preferences (dictionary): The preference holds the dictionary which have the value and keys.
        tieBreak (string): It holds the preference for the tiebreaker rule.

    Returns:
        [int]: the winner value.
    """
    tempdict={}
    for value in preferences.values():
        alternative = value[0]
        if alternative not in tempdict.keys():
            tempdict[alternative] = 1
        else:
            tempdict[alternative] += 1

    #print(tempdict)

    # get the list of highest scorers
    highestScores = []
    maxValue = max(tempdict.values())
    for key, value in tempdict.items():
        if value == maxValue:
            highestScores.append(key)
   
    # find the winner based on the tiebreak
    if tieBreak == 'max':
        return max(highestScores)
    elif tieBreak == 'min':
        return min(highestScores)
    else:
        agent = tieBreak
        temp_dict = {}
        for alternative in highestScores:
            temp_dict[alternative] = preferences[agent].index(alternative)
        # print(preferences[agent])
        # print(temp_dict)
    winner = min(temp_dict, key=temp_dict.get)
    return winner


def veto(preferences, tieBreak):
    """Every agent assigns 0 points to the alternative that they rank in the last place of their preference orderings, and 1 point to every other alternative.
     The winner is the alternative with the most number of points

    Args:
        preferences (dictionary): The preference holds the dictionary which have the value and keys.
        tieBreak (string): It holds the preference for the tiebreaker rule.

    Returns:
        [int]: the winner value.
    """
    Vetodict = {}
    for sortedList in preferences.values():
        for alternative in sortedList[:-1]:
            if alternative not in Vetodict.keys():
                Vetodict[alternative] = 1
            else:
                Vetodict[alternative] += 1
    #print(Vetodict)
    # get the list of highest scorers
    highestScores = []
    maxValue = max(Vetodict.values())
    for key, value in Vetodict.items():
        if value == maxValue:
            highestScores.append(key)
   
    # find the winner based on the tiebreak
    if tieBreak == 'max':
        return max(highestScores)
    elif tieBreak == 'min':
        return min(highestScores)
    else:
        agent = tieBreak
        temp_dict = {}
        for alternative in highestScores:
            temp_dict[alternative] = preferences[agent].index(alternative)
        # print(preferences[agent])
        # print(temp_dict)
    winner = min(temp_dict, key=temp_dict.get)
    return winner

def borda(preferences,tieBreak):
    """Every agent assigns a score of 0 to the their least-preferred alternative (the one at the bottom of the preference ranking), 
    a score of 1 to the second least-preferred alternative, ... , and a score of m-1 to their favourite alternative. 
    In other words, the alternative ranked at position j receives a score of m-j. The winner is the alternative with the highest score.

    Args:
        preferences (dictionary): The preference holds the dictionary which have the value and keys.
        tieBreak (string): It holds the preference for the tiebreaker rule.

    Returns:
        [int]: the winner value.
    """
    bordadict={}
    m=len(preferences[1])
    for alternative in preferences[1]:
        bordadict[alternative]=0
    for value in preferences.values():
          for index, alternative in enumerate(value):
                bordadict[alternative]+=m-(index+1)
    
    # get the list of highest scorers
    highestScores = []
    maxValue = max(bordadict.values())
    for key, value in bordadict.items():
        if value == maxValue:
            highestScores.append(key)
   
    # find the winner based on the tiebreak
    if tieBreak == 'max':
        return max(highestScores)
    elif tieBreak == 'min':
        return min(highestScores)
    else:
        agent = tieBreak
        temp_dict = {}
        for alternative in highestScores:
            temp_dict[alternative] = preferences[agent].index(alternative)
        # print(preferences[agent])
        # print(temp_dict)
    winner = min(temp_dict, key=temp_dict.get)
    return winner

def harmonic(preferences,tieBreak):
    """Every agent assigns a score of 1/m to the their least-preferred alternative (the one at the bottom of the preference ranking), 
    a score of 1/(m-1) to the second least-preferred alternative, ... , and a score of 1 to their favourite alternative. 
    In other words, the alternative ranked at position j receives a score of 1/j.

    Args:
        preferences (dictionary): The preference holds the dictionary which have the value and keys.
        tieBreak (string): It holds the preference for the tiebreaker rule.

    Returns:
        [int]: the winner value.
    """
    harmonicdict={}
    for alternative in preferences[1]:
        harmonicdict[alternative]=0
    for value in preferences.values():
          for index, alternative in enumerate(value):
                harmonicdict[alternative]+=1/(index+1)

    # get the list of highest scorers
    highestScores = []
    maxValue = max(harmonicdict.values())
    for key, value in harmonicdict.items():
        if value == maxValue:
            highestScores.append(key)
   
    # find the winner based on the tiebreak
    if tieBreak == 'max':
        return max(highestScores)
    elif tieBreak == 'min':
        return min(highestScores)
    else:
        agent = tieBreak
        temp_dict = {}
        for alternative in highestScores:
            temp_dict[alternative] = preferences[agent].index(alternative)
        # print(preferences[agent])
        # print(temp_dict)
    winner = min(temp_dict, key=temp_dict.get)
    return winner

def rangeVoting(values, tieBreak):
    """
    The worksheet is accessed from an xlsx file, and input values from the generatePrefence.
    The function return the alternative that has the maximum sum of valuations, 
    i.e., the maximum sum of numerical values in the xlsx file, using the tie-breaking option to distinguish between possible winners.

    Args:
        values (int): which holds the interger value
        tieBreak (string): It holds the preference for the tiebreaker rule.

    Returns:
        [int]: the winner value.
    """
    #values = workbook.active
    agents=values.max_row
    alternatives=values.max_column
    rangedict={}
    for column in range(1,(alternatives+1)):
        print('Entered column:', column)
        rangedict[column]=0
        print('Initialised rangedict {} with alternative {}'.format(rangedict, column))
        for row in range(1,(agents+1)):
            print("Entered row: {} in column: {}".format(row, column))
            CellValue=values.cell(row,column).value
            rangedict[column]+=CellValue
            print('Added cell value: {} to range dict {}'.format(CellValue, rangedict))

    # get the list of highest scorers
    highestScores = []
    preferences=generatePreferences(values)
    maxValue = max(rangedict.values())
    for key, value in rangedict.items():
        if value == maxValue:
            highestScores.append(key)
   
    # find the winner based on the tiebreak
    if tieBreak == 'max':
        return max(highestScores)
    elif tieBreak == 'min':
        return min(highestScores)
    else:
        agent = tieBreak
        temp_dict = {}
        for alternative in highestScores:
            temp_dict[alternative] = preferences[agent].index(alternative)
        # print(preferences[agent])
        # print(temp_dict)
    winner = min(temp_dict, key=temp_dict.get)
    return winner    

def scoringRule(preferences,scoreVector,tieBreak):
    """ For every agent, the function assigns the highest score in the scoring vector to the most preferred alternative of the agent, 
    the second highest score to the second most preferred alternative of the agent and so on, and the lowest score to the least preferred alternative of the agent. 
    In the end, it returns the alternative with the highest total score, using the tie-breaking option to distinguish between alternatives with the same score.

    Args:
        preferences (dictionary): The preference holds the dictionary which have the value and keys.
        scoreVector (int): 
        tieBreak (string): It holds the preference for the tiebreaker rule.

    Returns:
        [int]: the winner value.
    """
    ScoringDict={}
    scoreVector.sort(reverse=True)

    for value in preferences.values():
        for index,alternatives in enumerate(value):
            if alternatives not in ScoringDict.keys():
                ScoringDict[alternatives]=scoreVector[index]
            else:
                ScoringDict[alternatives]+=scoreVector[index]
    
    # get the list of highest scorers
    highestScores = []
    maxValue = max(ScoringDict.values())
    for key, value in ScoringDict.items():
        if value == maxValue:
            highestScores.append(key)
   
    # find the winner based on the tiebreak
    if tieBreak == 'max':
        return max(highestScores)
    elif tieBreak == 'min':
        return min(highestScores)
    else:
        agent = tieBreak
        temp_dict = {}
        for alternative in highestScores:
            temp_dict[alternative] = preferences[agent].index(alternative)
        # print(preferences[agent])
        # print(temp_dict)
    winner = min(temp_dict, key=temp_dict.get)
    return winner  

def STV(preferences, tieBreak):
    """ In Single Transferable Vote. In each round, the alternatives that appear the least frequently in the first position of agents' rankings are removed, 
    and the process is repeated. 
    When the final set of alternatives is removed (one or possibly more), then this last set is the set of possible winners.

    Args:
        preferences (dictionary): The preference holds the dictionary which have the value and keys.
        tieBreak (string): It holds the preference for the tiebreaker rule.

    Returns:
        [int]: the winner value.
    """
    temporary_preferences = copy.deepcopy(preferences)
    # temporary_preferences = preferences

    while True:

        # temporary dictionary for storing the frequency count of alternatives which resets to 0 after each updation
        alt_frequency = dict.fromkeys(temporary_preferences[1], 0)
        print('Temporary dictionary (for each round): ', temporary_preferences)
        for value in temporary_preferences.values():
            alt_frequency[value[0]] += 1
        print('Current alternative frequency (for each round):',alt_frequency)

        remove_alternatives = []
        min_value = min(alt_frequency.values())
        for key, value in alt_frequency.items():
            if value == min_value:
                remove_alternatives.append(key)
        # print(remove_alternatives)
        print('Least frequent alternatives to be removed: ', remove_alternatives, '\n')
        # get the final set of alternatives to be removed
        if len(remove_alternatives) == len(temporary_preferences[1]):
            print("Current temp preferences (first value): ", temporary_preferences[1])
            print("Don't remove this list of alternatives, as length equals the length of the current preferences list: ", remove_alternatives)
            print('\nWinner is:')
            return tiebreak(remove_alternatives, tieBreak, preferences)
        else:
            for alternative in remove_alternatives:
                alt_frequency.pop(alternative, None)

            # to update preferences values, by removing the least frequency keys(alternatives)
            for value in temporary_preferences.values():
                for alternative in remove_alternatives:
                    value.remove(alternative)
            



def tiebreak(list_of_winners, tieBreak, preferences):
    """ With the help of this function among the following possible winning alternatives, need to select the one with the following options:
    Max:The one with the highest number
    Min:The one with the lowest number
    agent i: The one that agent i ranks the highest in there preference ordering. 

    Args:
        list_of_winners (list): list of winner is the alternative preference for the winner based on tie breaker rule.
        tieBreak (string): It allows the user to select the option between max,min and agent option based on the prefence input.
        preferences (list): It holds the preferred value for the list

    Returns:
        [int]: it returns the min and max value for the tie breaker value.
    """
    if tieBreak == 'max':
        return max(list_of_winners)
    elif tieBreak == 'min':
        return min(list_of_winners)
    else:
        agent = tieBreak
        temp_dict = {}
        for alternative in list_of_winners:
            temp_dict[alternative] = preferences[agent].index(alternative)
        return min(temp_dict, key=temp_dict.get)